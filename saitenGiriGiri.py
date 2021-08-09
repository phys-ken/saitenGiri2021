import tkinter
from tkinter import messagebox
from PIL import Image, ImageTk  # 外部ライブラリ
import csv
import glob
import os
import sys
import shutil
import pathlib
import imghdr
import openpyxl

global canvas1
global img_resized
global img_tk
global Giri_cutter
global root
global topimg
global topfig

global RESIZE_RETIO  # 縮小倍率の規定
window_h = 700
window_w = int(window_h * 1.7)
fig_area_w = int(window_h * 1)

global qCnt
qCnt = 0


# 初回起動時にフォルダを展開する。
def initDir():
    os.makedirs("./setting/input", exist_ok=True)
    os.makedirs("./setting/output", exist_ok=True)
    f = open('setting/ini.csv', 'w')  # 既存でないファイル名を作成してください
    writer = csv.writer(f, lineterminator='\n')  # 行末は改行
    writer.writerow(["tag", "start_x", "start_y", "end_x", "end_y"])
    f.close()


# 与えられたフォルダの全てのファイル(フルパス)をソートしてリストで返す
# 拡張子が画像かどうかも判別し、画像のパスのみを返す。
def get_sorted_files(dir_path):
    all_sorted = sorted(glob.glob(dir_path))
    fig_sorted = [s for s in all_sorted if s.endswith(
        ('jpg', "jpeg", "png", "PNG", "JPEG", "JPG", "gif"))]
    return fig_sorted


# ドラッグ開始した時のイベント - - - - - - - - - - - - - - - - - - - - - - - - - -
def start_point_get(event):
    global start_x, start_y  # グローバル変数に書き込みを行なうため宣言

    canvas1.delete("rectTmp")  # すでに"rectTmp"タグの図形があれば削除

    # canvas1上に四角形を描画（rectangleは矩形の意味）
    canvas1.create_rectangle(event.x,
                             event.y,
                             event.x + 1,
                             event.y + 1,
                             outline="red",
                             tag="rectTmp")
    # グローバル変数に座標を格納
    start_x, start_y = event.x, event.y

# ドラッグ中のイベント - - - - - - - - - - - - - - - - - - - - - - - - - -


def rect_drawing(event):

    # ドラッグ中のマウスポインタが領域外に出た時の処理
    if event.x < 0:
        end_x = 0
    else:
        end_x = min(img_resized.width, event.x)
    if event.y < 0:
        end_y = 0
    else:
        end_y = min(img_resized.height, event.y)

    # "rectTmp"タグの画像を再描画
    canvas1.coords("rectTmp", start_x, start_y, end_x, end_y)

# ドラッグを離したときのイベント - - - - - - - - - - - - - - - - - - - - - - - - - -


def release_action(event):
    global qCnt

    if qCnt == 0:
        pos = canvas1.bbox("rectTmp")

        # canvas1上に四角形を描画（rectangleは矩形の意味）
        create_rectangle_alpha(pos[0], pos[1], pos[2], pos[3],
                               fill="green",
                               alpha=0.3,
                               tag="nameBox"
                               )

        canvas1.create_text(
            (pos[0] + pos[2]) / 2, (pos[1] + pos[3]) / 2,
            text="name",
            tag="nameText"
        )

        # "rectTmp"タグの画像の座標を元の縮尺に戻して取得
        start_x, start_y, end_x, end_y = [
            round(n * RESIZE_RETIO) for n in canvas1.coords("rectTmp")
        ]
        with open('setting/ini.csv', 'a') as f:
            writer = csv.writer(f, lineterminator='\n')  # 行末は改行
            writer.writerow(["name", start_x, start_y, end_x, end_y])

    else:
        pos = canvas1.bbox("rectTmp")
        # canvas1上に四角形を描画（rectangleは矩形の意味）
        create_rectangle_alpha(pos[0], pos[1], pos[2], pos[3],
                               fill="red",
                               alpha=0.3,
                               tag="qBox" + str(qCnt)
                               )
        canvas1.create_text(
            (pos[0] + pos[2]) / 2, (pos[1] + pos[3]) / 2,
            text="Q_" + str(qCnt),
            tag="qText" + str(qCnt)
        )

        # "rectTmp"タグの画像の座標を元の縮尺に戻して取得
        start_x, start_y, end_x, end_y = [
            round(n * RESIZE_RETIO) for n in canvas1.coords("rectTmp")
        ]
        with open('setting/ini.csv', 'a') as f:
            writer = csv.writer(f, lineterminator='\n')  # 行末は改行
            writer.writerow(["Q_" + str(qCnt).zfill(4),
                            start_x, start_y, end_x, end_y])

    qCnt = qCnt + 1


# 透過画像の作成 - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
# https://stackoverflow.com/questions/54637795/how-to-make-a-tkinter-canvas-rectangle-transparent/54645103
# 透過画像を削除するときは、imagesの配列から消す。
images = []  # to hold the newly created image


def create_rectangle_alpha(x1, y1, x2, y2, **kwargs):
    if 'alpha' in kwargs:
        alpha = int(kwargs.pop('alpha') * 255)
        fill = kwargs.pop('fill')
        fill = Giri_cutter.winfo_rgb(fill) + (alpha,)
        image = Image.new('RGBA', (x2-x1, y2-y1), fill)
        images.append(ImageTk.PhotoImage(image, master=Giri_cutter))
        canvas1.create_image(x1, y1, image=images[-1], anchor='nw')
    canvas1.create_rectangle(x1, y1, x2, y2, **kwargs)


def back_one():
    global qCnt
    if qCnt == 0:
        return
    qCnt = qCnt - 1
    # タグに基づいて画像を削除
    if qCnt == 0:
        canvas1.delete("nameBox", "nameText", "rectTmp")
        images.pop(-1)
    else:
        canvas1.delete("qBox" + str(qCnt), "qText" + str(qCnt), "rectTmp")
        images.pop(-1)
    # csvの最終行を削除
    readFile = open("setting/ini.csv")
    lines = readFile.readlines()
    readFile.close()
    w = open("setting/ini.csv", 'w')
    w.writelines([item for item in lines[:-1]])
    w.close()


def trim_fin():
    global Giri_cutter
    ret = messagebox.askyesno('終了します', '斬り方を決定し、ホームに戻っても良いですか？')
    if ret == True:
        cur = os.getcwd()
        beforePath = cur + "/setting/ini.csv"
        afterPath = cur + "/setting/trimData.csv"
        shutil.move(beforePath, afterPath)
        Giri_cutter.destroy()


def setting_ck():
    if not os.path.exists("./setting/"):
        ret = messagebox.askyesno(
            '初回起動です', '採点のために、いくつかのフォルダーをこのファイルと同じ場所に作成します。\nよろしいですか？')
        if ret == True:
            initDir()
            messagebox.showinfo(
                '準備ができました。', '解答用紙を、setting/input の中に保存してください。jpeg または png に対応しています。')
        else:
            # メッセージボックス（情報）
            messagebox.showinfo('終了', 'フォルダは作成しません。')
    else:
        messagebox.showinfo(
            '確認', '初期設定は完了しています。解答用紙を、setting/inputに入れてから、解答用紙分割をしてください。')


def input_ck():
    # 表示する画像の取得
    files = get_sorted_files(os.getcwd() + "/setting/input/*")
    if not files:
        # メッセージボックス（警告）
        messagebox.showerror(
            "エラー", "setting/inputの中に、解答用紙のデータが存在しません。画像を入れてから、また開いてね。")
    else:
        GiriActivate()


def GiriActivate():
    global RESIZE_RETIO
    global img_resized
    global canvas1
    global img_tk
    global Giri_cutter
    global qCnt

    def toTop():
        global qCnt
        ret = messagebox.askyesno(
            '保存しません', '作業中のデータは保存されません。\n画面を移動しますか？')
        if ret == True:
            qCnt = 0
            Giri_cutter.destroy()
        else:
            pass

    # 表示する画像の取得
    files = get_sorted_files(os.getcwd() + "/setting/input/*")
    print(files)

    # ini.csvは、起動のたびに初期化する。
    f = open('setting/ini.csv', 'w')  # 既存でないファイル名を作成してください
    writer = csv.writer(f, lineterminator='\n')  # 行末は改行
    writer.writerow(["tag", "start_x", "start_y", "end_x", "end_y"])
    f.close()

    img = Image.open(files[0])

    # 画面サイズに合わせて画像をリサイズする
    # 画像サイズが縦か横かに合わせて、RESIZE_RETIOを決める。
    w, h = img.size
    if w >= h:
        if w <= fig_area_w:
            RESIZE_RETIO = 1
        else:
            RESIZE_RETIO = w / fig_area_w
    else:
        if h <= window_h:
            RESIZE_RETIO = 1
        else:
            RESIZE_RETIO = h / window_h

    # 画像リサイズ
    img_resized = img.resize(size=(int(img.width / RESIZE_RETIO),
                                   int(img.height / RESIZE_RETIO)),
                             resample=Image.BILINEAR)

    Giri_cutter = tkinter.Tk()
    Giri_cutter.geometry(str(window_w) + "x" + str(window_h))
    Giri_cutter.title("解答用紙を斬る")

    cutting_frame = tkinter.Frame(Giri_cutter)
    cutting_frame.pack()
    canvas_frame = tkinter.Frame(cutting_frame)
    canvas_frame.grid(column=0, row=0)
    button_frame = tkinter.Frame(cutting_frame)
    button_frame.grid(column=1, row=0)

    # tkinterで表示できるように画像変換
    img_tk = ImageTk.PhotoImage(img_resized, master=Giri_cutter)

    # Canvasウィジェットの描画
    canvas1 = tkinter.Canvas(canvas_frame,
                             bg="black",
                             width=img_resized.width,
                             height=img_resized.height,
                             highlightthickness=0)
    # Canvasウィジェットに取得した画像を描画
    canvas1.create_image(0, 0, image=img_tk, anchor=tkinter.NW)

    # Canvasウィジェットを配置し、各種イベントを設定
    canvas1.pack()

    # 戻るボタン
    backB = tkinter.Button(
        button_frame, text='一つ前に戻る', command=back_one, width=20, height=4).pack()

    # 入力完了
    finB = tkinter.Button(
        button_frame, text='入力完了\n(保存して戻る)', command=trim_fin, width=20, height=4).pack()
    topB = tkinter.Button(
        button_frame, text='topに戻る\n(保存はされません)', command=toTop, width=20, height=4).pack()

    canvas1.bind("<ButtonPress-1>", start_point_get)
    canvas1.bind("<Button1-Motion>", rect_drawing)
    canvas1.bind("<ButtonRelease-1>", release_action)
    Giri_cutter.mainloop()


def trimck():
    ret = messagebox.askyesno(
        'すべての解答用紙を斬っちゃいます。', '全員の解答用紙を、斬ります。\n以下の注意を読んで、よければ始めてください。\n①10分くらい時間がかかります。\n②inputに保存された画像は、削除されません。\n③現在のoutputは全て消えます。')
    if ret == True:
        allTrim()
    else:
        pass


def allTrim():
    # トリミング前の画像の格納先
    ORIGINAL_FILE_DIR = "./setting/input"
    # トリミング後の画像の格納先
    TRIMMED_FILE_DIR = "./setting/output"

    def trim(path, left, top, right, bottom):
        im = Image.open(path)
        im_trimmed = im.crop((left, top, right, bottom))
        return im_trimmed

    def readCSV():
        # もしcsvが無ければ、全部止める
        if os.path.isfile("./setting/trimData.csv") == False:
            return 0
        else:
            with open('./setting/trimData.csv') as f:
                reader = csv.reader(f)
                data = [row for row in reader]
                data.pop(0)
                return data

    data = readCSV()

    try:
        shutil.rmtree("./setting/output")
    except OSError as err:
        pass

    if data == 0:
        messagebox.showinfo('終了', 'どうやって斬ればいいかわかりません。\nまずはどこを斬るかを決めてください。')
        return 0

    while data:
        title, left, top, right, bottom = data.pop(0)
        print(title)
        print(left, top, right, bottom)

        outputDir = TRIMMED_FILE_DIR + "/" + title

        # もしトリミング後の画像の格納先が存在しなければ作る
        if os.path.isdir(outputDir) == False:
            os.makedirs(outputDir)

        # 画像ファイル名を取得
        files = os.listdir(ORIGINAL_FILE_DIR)
        # 特定の拡張子のファイルだけを採用。実際に加工するファイルの拡張子に合わせる
        files = [name for name in files if name.split(
            ".")[-1] in ['jpg', "jpeg", "png", "PNG", "JPEG", "JPG", "gif"]]

        try:
            for val in files:
                # オリジナル画像へのパス
                path = ORIGINAL_FILE_DIR + "/" + val
                # トリミングされたimageオブジェクトを取得
                im_trimmed = trim(path, int(left), int(top),
                                  int(right), int(bottom))
                # トリミング後のディレクトリに保存。ファイル名の頭に"cut_"をつけている
                # qualityは95より大きい値は推奨されていないらしい
                im_trimmed.save(outputDir + "/" + val, quality=95)

            print("トリミングが終了しました。")
            print("********************************")
        except:
            messagebox.showinfo(
                'エラー', 'エラーが検出されました。中断します。\n\n' + str(sys.stderr))
            try:
                shutil.rmtree("./setting/output")
            except OSError as err:
                pass
            return 0
    output_Sh()
    messagebox.showinfo('斬りました', '全員分の解答用紙を斬りました。')


def exitGiri():
    sys.exit()


def info():
    pass


def saiten2xlsx():
    def readCSV():
        # もしcsvが無ければ、全部止める
        if os.path.isfile("./setting/trimData.csv") == False:
            return 0
        else:
            with open('./setting/trimData.csv') as f:
                reader = csv.reader(f)
                data = [row for row in reader]
                data.pop(0)
                return data

    def setTensu(figname, qname, tensu):
        xlPath = "./setting/saiten.xlsx"
        #wb = openpyxl.load_workbook(xlPath)
        #ws = wb["採点シート"]

        qCol = int(qname[-4:]) + 3
        ws.cell(1, qCol + 1).value = qname

        MIN_COL = 1
        MIN_ROW = 2

        MAX_COL = 1
        MAX_ROW = ws.max_row

        # 範囲データを順次処理
        for row in ws.iter_rows(min_col=MIN_COL, min_row=MIN_ROW, max_col=MAX_COL, max_row=MAX_ROW):
            for cell in row:
                try:
                    # 該当セルの値取得
                    cell_value = cell.value
                    if figname == cell_value:
                        o = cell.offset(0, qCol)
                        try:
                            o.value = int(tensu)
                        except:
                            o.value = tensu
                except:
                    pass

        #wb.save(xlPath)

    data = readCSV()

    xlPath = "./setting/saiten.xlsx"
    wb = openpyxl.load_workbook(xlPath)
    ws = wb["採点シート"]

    while data:
        title, left, top, right, bottom = data.pop(0)
        if title == "name":
            continue
        qpath = "./setting/output/" + title
        for curDir, dirs, files in os.walk(qpath):
            if files:
                for f in files:
                    tensu = os.path.basename(os.path.dirname(curDir + "/" + f))
                    if not tensu == title:
                        setTensu(figname=f, qname=title, tensu=tensu)
                    else:
                        setTensu(figname=f, qname=title, tensu="未")
    wb.save(xlPath)
    


def saitenSelect():
    def show_selection():
        for i in lb.curselection():
            print(lb.get(i))
            siwakeApp(str(lb.get(i)))
            selectQ.destroy()

    def backTop():
        selectQ.destroy()

    selectQ = tkinter.Tk()
    selectQ.geometry("500x500")
    selectQ.title("採点する問題を選ぶ")

    # outputの中のフォルダを取得
    path = "./setting/output/"
    files = os.listdir(path)
    files_dir = [f for f in files if os.path.isdir(os.path.join(path, f))]
    files_dir.sort()
    lb = tkinter.Listbox(selectQ, selectmode='single', height=20, width=20)
    for i in files_dir:
        if not i == "name":
            lb.insert(tkinter.END, i)

    lb.grid(row=0, column=0)
    button_frame = tkinter.Frame(selectQ)
    button_frame.grid(row=0, column=1)

    button1 = tkinter.Button(
        button_frame, text='採点する', width=15, height=3,
        command=lambda: show_selection()).pack()

    totopB = tkinter.Button(
        button_frame, text='Topに戻る', width=15, height=3,
        command=backTop).pack()

    selectQ.mainloop


def folder_walker(folder_path, recursive=False, file_ext=".*"):
    """
    指定されたフォルダのファイル一覧を取得する。
    引数を指定することで再帰的にも、非再帰的にも取得可能。

    Parameters
    ----------
    folder_path : str
        対象のフォルダパス
    recursive : bool
        再帰的に取得するか否か。既定値はTrueで再帰的に取得する。
    file_ext : str
        読み込むファイルの拡張子を指定。例：".jpg"のようにピリオドが必要。既定値は".*"で指定なし
    """

    p = pathlib.Path(folder_path)

    if recursive:
        return list(p.glob("**/*" + file_ext))  # **/*で再帰的にファイルを取得
    else:
        return list(p.glob("*" + file_ext))  # 再帰的にファイル取得しない


# ファイル読み込み - - - - - - - - - - - - - - - - - - - - - - - -
def load_file(Qnum):

    global img_num, item, dir_name

    # ファイルを読み込み
    tex_var.set("ファイルを読み込んでいます...")
    dir_name = "./setting/output/" + Qnum
    if not dir_name == None:
        file_list = folder_walker(dir_name)

    # ファイルから読み込める画像をリストに列挙
    for f in file_list:
        try:
            img_lst.append(Image.open(f))
            filename_lst.append(f)
        except:
            pass

    if not img_lst:
        tex_var.set("読み込む画像がありません。\n採点は終了しています。")

    # ウィンドウサイズに合わせてキャンバスサイズを再定義
    # window_resize()

    # 画像変換
    for f in img_lst:

        # キャンバス内に収まるようリサイズ
        resized_img = img_resize_for_canvas(f, image_canvas)

        # tkinterで表示できるように画像変換
        tk_img_lst.append(ImageTk.PhotoImage(
            image=resized_img, master=image_canvas))

    # キャンバスの中心を取得
    c_width_half = round(int(image_canvas["width"]) / 2)
    c_height_half = round(int(image_canvas["height"]) / 2)

    # キャンバスに表示
    img_num = 0
    item = image_canvas.create_image(
        c_width_half, c_height_half,  image=tk_img_lst[0], anchor=tkinter.CENTER)
    # ラベルの書き換え
    tex_var.set(filename_lst[img_num])

    # 仕分け実行ボタンの配置
    assort_btn.pack()

# 次の画像へ - - - - - - - - - - - - - - - - - - - - - - - -


def next_img(event):
    global img_num

    # 読み込んでいる画像の数を取得
    img_count = len(tk_img_lst)

    # 画像が最後でないか判定
    if img_num >= img_count - 1:
        pass
    else:
        # 表示中の画像No.を更新して表示
        img_num += 1
        image_canvas.itemconfig(item, image=tk_img_lst[img_num])
        # ラベルの書き換え
        tex_var.set(filename_lst[img_num])
        # ラベリングを表示
        if filename_lst[img_num] in assort_dict:
            assort_t_var.set(assort_dict[filename_lst[img_num]])
        else:
            assort_t_var.set("")


# 前の画像へ - - - - - - - - - - - - - - - - - - - - - - - -
def prev_img(event):
    global img_num

    # 画像が最初でないか判定
    if img_num <= 0:
        pass
    else:
        # 表示中の画像No.を更新して表示
        img_num -= 1
        image_canvas.itemconfig(item, image=tk_img_lst[img_num])
        # ラベルの書き換え
        tex_var.set(filename_lst[img_num])
        # ラベリングを表示
        if filename_lst[img_num] in assort_dict:
            assort_t_var.set(assort_dict[filename_lst[img_num]])
        else:
            assort_t_var.set("")


# ウィンドウサイズからキャンバスサイズを再定義 - - - - - - - - - - - - - - - - -
def window_resize():

    image_canvas["width"] = image_canvas.winfo_width()
    image_canvas["height"] = image_canvas.winfo_height()


# キャンバスサイズに合わせて画像を縮小 - - - - - - - - - - - - - - - - - - - -
def img_resize_for_canvas(img, canvas, expand=False):

    size_retio_w = int(canvas["width"]) / img.width
    size_retio_h = int(canvas["height"]) / img.height

    if expand == True:
        size_retio = min(size_retio_w, size_retio_h)
    else:
        size_retio = min(size_retio_w, size_retio_h, 1)

    resized_img = img.resize((round(img.width * size_retio),
                              round(img.height * size_retio)))
    return resized_img

# 画像表示 - - - - - - - - - - - - - - - - - - - - - - - -


def image_show(event):
    img_lst[img_num].show()


# 画像に対しラベリング - - - - - - - - - - - - - - - - - - - - - - - -
def file_assort(event):

    if str(event.keysym) in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
        assort_dict[filename_lst[img_num]] = str(event.keysym)
    else:
        assort_dict[filename_lst[img_num]] = str("skip")

    # ラベリングを表示
    if filename_lst[img_num] in assort_dict:
        assort_t_var.set(assort_dict[filename_lst[img_num]])
    else:
        assort_t_var.set("")

    print(assort_dict[filename_lst[img_num]])


# フォルダ分け実行 - - - - - - - - - - - - - - - - - - - - - - - -
def assort_go(event):

    global f_dir

    for f in assort_dict:
        # 仕分け前後のファイル名・フォルダ名を取得
        # skipは、無視する。
        print(assort_dict[f])
        if not assort_dict[f] == "skip":
            f_dir = os.path.dirname(f)
            f_basename = os.path.basename(f)
            new_dir = os.path.join(f_dir, assort_dict[f])
            new_path = os.path.join(new_dir, f_basename)

            # ディレクトリの存在チェック
            if not os.path.exists(new_dir):
                os.mkdir(new_dir)
            # ファイルの移動実行
            shutil.move(f, new_path)

            print(new_path)
        else:
            pass
    saiten2xlsx()
    messagebox.showinfo("採点保存", "ここまでの採点結果を保存しました。\nskipした項目は、採点されていません。")
    siwake_win.destroy()


def siwakeApp(Qnum):
    def exit_siwake():
        ret = messagebox.askyesno('終了します', '採点を中断し、ホームに戻っても良いですか？')
        if ret == True:
            siwake_win.destroy()

    # グローバル変数
    global img_lst, tk_img_lst
    global filename_lst
    global assort_file_list
    global assort_dict

    global tex_var
    global image_canvas
    global assort_btn
    global assort_t_var

    global img_num
    global f_basename
    global siwake_win

    img_lst, tk_img_lst = [], []
    filename_lst = []
    assort_file_list = []
    assort_dict = {}

    print("siwakeApp IN " + str(Qnum))

    img_num = 0
    f_basename = ""

    siwake_win = tkinter.Tk()
    siwake_win.title("採点中...")
    siwake_win.geometry("800x800")
    siwake_frame = tkinter.Frame(siwake_win)
    siwake_frame.grid(column=0, row=0)
    button_siwake_frame = tkinter.Frame(siwake_win)
    button_siwake_frame.grid(column=1, row=0)

    # キャンバス描画設定
    image_canvas = tkinter.Canvas(siwake_frame,
                                  width=640,
                                  height=480)

    image_canvas.pack(expand=True, fill="both")

    # 仕分け結果表示
    assort_t_var = tkinter.StringVar(siwake_frame)
    assort_label = tkinter.Label(
        siwake_frame, textvariable=assort_t_var, font=("Meiryo UI", 30))
    assort_label.pack()

    # ファイル名ラベル描画設定
    tex_var = tkinter.StringVar(siwake_frame)
    tex_var.set("ファイル名")

    lbl = tkinter.Label(siwake_frame, textvariable=tex_var,
                        font=("Meiryo UI", 20))
    lbl["foreground"] = "gray"
    lbl.pack()

    # 右左キーで画像送りする動作設定
    siwake_win.bind("<Key-Right>", next_img)
    siwake_win.bind("<Key-Left>", prev_img)
    # 「Ctrl」+「P」で画像表示
    siwake_win.bind("<Control-Key-p>", image_show)

    # 数字キーで仕分け対象設定
    siwake_win.bind("<Key>", file_assort)

    # 仕分け実行ボタン
    assort_btn = tkinter.Button(
        button_siwake_frame, text="採点実行",  height=3, width=15)
    assort_btn.bind("<Button-1>", assort_go)

    exit_button = tkinter.Button(
        button_siwake_frame, text="トップに戻る\n保存はされません", height=3, width=15,  command=exit_siwake)
    exit_button.pack()

    # 読み込みボタン描画設定
    load_file(Qnum)

    siwake_win.mainloop


def output_Sh():

    # 定数設定
    SHEET_TITLE = '採点シート'  # シート名の設定
    RESULT_FILE_NAME = './setting/saiten.xlsx'  # 結果を保存するファイル名

    # 変数
    max_height = []  # 各行の画像の高さの最大値を保持

    def get_file_names(set_dir_name):
        """
        ディレクトリ内のファイル名取得（ファイル名のみの一覧を取得）
        """
        file_names = os.listdir(set_dir_name)
        temp_full_file_names = [os.path.join(set_dir_name, file_name) for file_name in file_names if os.path.isfile(
            os.path.join(set_dir_name, file_name))]  # ファイルかどうかを判定
        return temp_full_file_names

    def attach_img(target_full_file_names, set_column_idx, set_dir_name):
        """
        画像を呼び出して、Excelに貼り付け
        """
        set_row_idx = 1
        column_letter = "B"
        # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
        ws.cell(row=1, column=set_column_idx).value = "画像"
        ws.cell(row=1, column=1).value = "ファイル名"  # ファイル名
        ws.cell(row=1, column=3).value = "生徒番号"  # 出席番号
        ws.cell(row=1, column=4).value = "名前"  # 名前
        max_width = 0  # 画像の幅の最大値を保持するための変数
        target_full_file_names.sort()  # ファイル名でソート
        for target_file in target_full_file_names:
            p = pathlib.Path(target_file)
            target_file = p.resolve()
            if imghdr.what(target_file) != None:  # 画像ファイルかどうかの判定
                img = openpyxl.drawing.image.Image(target_file)
                #print('[' + column_letter + '][' + str(set_row_idx+1) + ']' + target_file + 'を貼り付け')

                # 画像のサイズを取得して、セルの大きさを合わせる（画像同士が重ならないようにするため）
                size_img = Image.open(target_file)
                width, height = size_img.size
                if max_width < width:
                    max_width = width
                # 配列「max_height」において、「set_row_idx」番目の要素が存在しなければ、挿入
                if not max_height[set_row_idx-1:set_row_idx]:
                    max_height.insert(set_row_idx-1, height)
                if max_height[set_row_idx-1] < height:
                    max_height[set_row_idx-1] = height
                ws.row_dimensions[set_row_idx +
                                  1].height = max_height[set_row_idx-1] * 0.75
                ws.column_dimensions[column_letter].width = int(
                    max_width) * 0.13

                # セルの行列番号から、そのセルの番地を取得
                cell_address = ws.cell(
                    row=set_row_idx + 1, column=set_column_idx).coordinate
                img.anchor = cell_address
                ws.add_image(img)  # シートに画像貼り付け
                ws.cell(row=set_row_idx + 1,
                        column=1).value = os.path.basename(target_file)

            set_row_idx += 1

    # ワークブック設定
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]  # 1番目のシートを編集対象にする
    ws.title = SHEET_TITLE  # 1番目のシートに名前を設定

    # 貼り付ける画像を置いておくルートディレクトリ内のディレクトリ名を再帰的に取得
    dir_name = "./setting/output/name"

    column_idx = 2

    f_names = get_file_names(dir_name)  # ファイル名取得
    print(f_names)
    attach_img(f_names, column_idx, dir_name)  # 画像貼り付け設定

    # ファイルへの書き込み
    wb.save(RESULT_FILE_NAME)


def top_activate():

    val = 0.4
    fifwid = 500
    fifhet = 400
    global root
    global topimg
    global topfig

    global top_frame
    top_frame = tkinter.Frame(root, bg="white")
    top_frame.pack()
    fig_frame = tkinter.Frame(top_frame, width=fifwid, height=fifhet)
    fig_frame.grid(column=0, row=0)

    topimg = Image.open("./appfigs/top.png")
    topimg = topimg.resize(
        (int(topimg.width * val), int(topimg.height * val)), 0)
    topfig = ImageTk.PhotoImage(topimg, master=root)
    canvas_top = tkinter.Canvas(
        bg="white", master=fig_frame, width=fifwid + 30, height=fifhet, highlightthickness=0)
    canvas_top.place(x=0, y=0)
    canvas_top.create_image(0, 0, image=topfig, anchor=tkinter.NW)
    canvas_top.pack()

    button_frame = tkinter.Frame(top_frame, bg="white", highlightthickness=0)
    button_frame.grid(column=1, row=0)

    infoB = tkinter.Button(
        button_frame, text="はじめに", command=info, width=15, height=2, highlightthickness=0).pack()

    initB = tkinter.Button(
        button_frame, text="初期設定をする", command=setting_ck, width=15, height=2, highlightthickness=0).pack()

    GiriGoB = tkinter.Button(
        button_frame, text="どこを斬るか決める", command=input_ck, width=15, height=2, highlightthickness=0).pack()

    initB = tkinter.Button(
        button_frame, text="全員の解答用紙を斬る", command=trimck, width=15, height=2, highlightthickness=0).pack()

    saitenB = tkinter.Button(
        button_frame, text="斬った画像を採点する", command=saitenSelect, width=15, height=2, highlightthickness=0).pack()

    exitB = tkinter.Button(
        button_frame, text="アプリを閉じる", command=exitGiri, width=15, height=2, highlightthickness=0).pack()


# メイン処理 - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
if __name__ == "__main__":

    # 画面処理
    root = tkinter.Tk()
    root.title("採点ギリギリ")
    root.geometry("800x400")
    root.configure(bg='white')

    top_activate()

    root.mainloop()
