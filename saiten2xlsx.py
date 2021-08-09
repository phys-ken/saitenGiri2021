import csv
import os
import openpyxl

def readCSV():
        # もしcsvが無ければ、全部止める
        if os.path.isfile("./setting/trimData.csv") == False:
            return 0
        else:
            with open('./setting/trimData.csv') as f:
                reader = csv.reader(f)
                data = [row for row in reader]
                data.pop(0)
                return data

def setTensu(figname , qname , tensu):
  xlPath = "./setting/saiten.xlsx"
  wb = openpyxl.load_workbook(xlPath)
  ws = wb["採点シート"]

  qCol = int(qname[-4:]) + 3
  ws.cell(1,qCol + 1).value = qname

  MIN_COL = 1
  MIN_ROW = 2
  
  MAX_COL = 1
  MAX_ROW = ws.max_row
  
  # 範囲データを順次処理
  for row in ws.iter_rows(min_col=MIN_COL, min_row=MIN_ROW, max_col=MAX_COL, max_row=MAX_ROW ):
      for cell in row:
        try:
          # 該当セルの値取得
          cell_value = cell.value
          if figname == cell_value:
            o = cell.offset(0,qCol)
            try:
              o.value = int(tensu)
            except:
              o.value = tensu


        except:
          pass
          
  wb.save(xlPath)



data = readCSV()

xlPath = "./setting/saiten.xlsx"
wb = openpyxl.load_workbook(xlPath)
ws = wb["採点シート"]



while data:
  title, left, top, right, bottom = data.pop(0)
  if  title == "name":
    continue
  qpath = "./setting/output/" + title
  print(title)
  print("Path = " +  qpath)

  for curDir, dirs, files in os.walk(qpath):
    if files:
      for f in files:
        tensu = os.path.basename(os.path.dirname(curDir + "/"+ f))
        if not tensu == title:
          print(f + "の得点は、" + str(tensu) + "点です。")
          setTensu(figname=f , qname = title , tensu=tensu)
        else:
          setTensu(figname=f , qname = title , tensu="未")


