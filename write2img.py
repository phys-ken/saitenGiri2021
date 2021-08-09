from PIL import Image, ImageDraw, ImageFont
import glob
import os
import csv
import openpyxl
import shutil

def writeImg():
  def readCSV_loc():
      # もしcsvが無ければ、全部止める
      if os.path.isfile("./setting/trimData.csv") == False:
          return 0
      else:
          with open('./setting/trimData.csv') as f:
              reader = csv.reader(f)
              data = [row for row in reader]
              data.pop(0)
              return data

  def getPoint(figName , qName ):
      qCol = int(qName[-4:]) + 3
      xlPath = "./setting/saiten.xlsx"
      wb = openpyxl.load_workbook(xlPath)
      ws = wb["採点シート"]
      list = [[0 for i in range(ws.max_column)] for j in range(ws.max_row)]
      for x in range(0, ws.max_row):
        for y in range(0, ws.max_column):
          list[x][y] = ws.cell(row=x+1, column=y+1).value
      for row in list:
          if figName in row:
              result = True
              break
      result

      if not str(row[qCol]) == "未":
        return str(row[qCol])
      else:
        return "?"

  files = glob.glob("./setting/input/*")
  data = readCSV_loc()
  img = Image.open(files[0])
  draw = ImageDraw.Draw(img)  # ImageDrawオブジェクトを作成

  if data:
    try:
      shutil.rmtree("./setting/kaitoYousi")
    except:
      print("kaitoYousiフォルダがまだありません、今から作ります。")
    os.makedirs("./setting/kaitoYousi", exist_ok=True)
    nmtitle, nmleft, nmtop, nmright, nmbottom = data.pop(0)
    h = abs(int(nmbottom) - int(nmtop))
    w = abs(int(nmright) - int(nmleft))
    if nmtitle == "name":
      if h >= w:
        mojiSize = int(w/2)
      else:
        mojiSize = int(h/2)
      font = ImageFont.truetype("arial.ttf", int(mojiSize))  # フォントを指定、64はサイズでピクセル単位

  for f in files:
    img = Image.open(f)
    draw = ImageDraw.Draw(img)  # ImageDrawオブジェクトを作成
    print(os.path.basename(f) + "を採点します。")
    sumVal = 0
    for pos in data:
      title, left, top, right, bottom = pos
      figName = os.path.basename(f)
      qName = title

      text = getPoint(figName , qName  ) 
      try:
        sumVal = sumVal + int(text)
      except:
        sumVal = sumVal

      # 画像
      draw.text((int(int(right) - mojiSize/2)  , int(top)), text, font=font , fill = "red")
      draw.rectangle((int(int(right) - mojiSize/2) , int(top),int(int(right) + mojiSize/2)  , int(top) + mojiSize ),  outline = "red" )

    
    #合計点を描画
    text = str(sumVal)
    draw.text((int(int(nmright) - mojiSize/2)  , int(nmtop)), text, font=font , fill = "red")
    draw.rectangle((int(int(nmright) - mojiSize/2) , int(nmtop),int(int(nmright) + mojiSize/2)  , int(nmtop) + mojiSize ),  outline = "red" )
    # ファイルを保存
    img.save('setting/kaitoYousi/' + os.path.basename(f) , quality=100, optimize=True)
    print(os.path.basename(f) + "の採点は正しく終了しました。")