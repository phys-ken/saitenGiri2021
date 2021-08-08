"""
Excelに画像貼り付け.py
"""
import os
import imghdr
import pathlib
import openpyxl
from PIL import Image, ImageTk  # 外部ライブラリ

def output_Sh():

  # 定数設定
  SHEET_TITLE = '画像貼り付け' # シート名の設定
  RESULT_FILE_NAME = './setting/name_register.xlsx' # 結果を保存するファイル名

  # 変数
  max_height = [] # 各行の画像の高さの最大値を保持


  def get_file_names(set_dir_name):
      """
      ディレクトリ内のファイル名取得（ファイル名のみの一覧を取得）
      """
      file_names = os.listdir(set_dir_name)
      temp_full_file_names = [os.path.join(set_dir_name, file_name) for file_name in file_names if os.path.isfile(os.path.join(set_dir_name, file_name))] # ファイルかどうかを判定
      return temp_full_file_names

  def attach_img(target_full_file_names, set_column_idx, set_dir_name):
      """
      画像を呼び出して、Excelに貼り付け
      """
      set_row_idx = 1
      column_letter = "B"
      ws.cell(row=1, column=set_column_idx).value = "画像" # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
      ws.cell(row=1, column=1).value = "ファイル名" # ファイル名
      ws.cell(row=1, column=3).value = "生徒番号" # 出席番号
      ws.cell(row=1, column=4).value = "名前" # 名前    
      max_width = 0 # 画像の幅の最大値を保持するための変数
      target_full_file_names.sort() # ファイル名でソート
      for target_file in target_full_file_names:
          p = pathlib.Path(target_file)
          target_file = p.resolve()
          if imghdr.what(target_file) != None: # 画像ファイルかどうかの判定
              img = openpyxl.drawing.image.Image(target_file)
              #print('[' + column_letter + '][' + str(set_row_idx+1) + ']' + target_file + 'を貼り付け')

              # 画像のサイズを取得して、セルの大きさを合わせる（画像同士が重ならないようにするため）
              size_img = Image.open(target_file)
              width , height = size_img.size
              if max_width < width:
                  max_width = width
              if not max_height[set_row_idx-1:set_row_idx]: # 配列「max_height」において、「set_row_idx」番目の要素が存在しなければ、挿入
                  max_height.insert(set_row_idx-1, height)
              if max_height[set_row_idx-1] < height:
                  max_height[set_row_idx-1] = height
              ws.row_dimensions[set_row_idx+1].height = max_height[set_row_idx-1] * 0.75
              ws.column_dimensions[column_letter].width = int(max_width) * 0.13

              cell_address = ws.cell(row=set_row_idx + 1, column=set_column_idx).coordinate # セルの行列番号から、そのセルの番地を取得
              img.anchor = cell_address
              ws.add_image(img) # シートに画像貼り付け
              ws.cell(row=set_row_idx + 1 , column = "A").value = os.path.basename(target_full_file_names)

          set_row_idx += 1


  # ワークブック設定
  wb = openpyxl.Workbook()
  ws = wb.worksheets[0] # 1番目のシートを編集対象にする
  ws.title = SHEET_TITLE # 1番目のシートに名前を設定

  # 貼り付ける画像を置いておくルートディレクトリ内のディレクトリ名を再帰的に取得
  dir_name = "./setting/output/name"

  column_idx = 2

  f_names = get_file_names(dir_name) # ファイル名取得
  print(f_names)
  attach_img(f_names, column_idx, dir_name) # 画像貼り付け設定

  # ファイルへの書き込み
  wb.save(RESULT_FILE_NAME)